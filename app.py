# -*- coding: utf-8 -*-
"""
Created on Sun Jun 14 13:51:07 2020

@author: amit
"""

import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import xlrd




DATA_URL = 'Book1.xlsx'

@st.cache(persist=True)

def load_data(nrows):
    data = pd.read_excel(DATA_URL, nrows=nrows)
    data.dropna(subset = ['latitude','longitude'],inplace=True)
    return data


user_input = st.text_input("Confirm your role (HOSPITAL/USER)")




if user_input=="USER":
    data = load_data(500)
    midpoint = (np.average(data['latitude']),np.average(data['longitude']))
    injured_people = st.slider("No. of beds available",0,2000)
    st.map(data.query("beds >= @injured_people")[["latitude","longitude"]].dropna(how="any"))
    data[data['beds']>injured_people][['hosp','beds']]
    midpoint = (np.average(data['latitude']),np.average(data['longitude']))
    if st.checkbox("Show Hospitals containing Ventilators",False):
        data[data['icu']>40][['hosp']]
    
    

elif user_input=="HOSPITAL":
    user_input5 = st.text_input("Enter Name of Hospital")
    user_input1 = st.text_input("Number of beds available in your hospital")
    user_input2 = st.text_input("Enter Latitude of your hospital (click button below to know)")
    user_input3 = st.text_input("Enter Longitude of your hospital (click button below to know)")
    user_input4 = st.text_input("Enter 100 if Ventilators available (else 0)")
    if user_input1!="" and user_input2!="" and user_input3!="" and user_input4!="" and user_input5!="":
        df = pd.DataFrame({'latitude':[user_input2],'longitude':[user_input3],'beds':[user_input1],'icu':[user_input4],'hosp':[user_input5]})
        book = load_workbook('Book1.xlsx')
        writer = pd.ExcelWriter('Book1.xlsx', engine='openpyxl')
        writer.book = book
        for sheetname in writer.sheets:
            df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
        writer.save()
        st.header("Value Entered Successfully")
    else:
        st.header("Enter the details")
    
    
    
    
    ##with pd.ExcelWriter("data.xlsx",
      ##              mode='a',engine="openpyxl") as writer:  
        ##df.to_excel(writer, sheet_name='Sheet1')